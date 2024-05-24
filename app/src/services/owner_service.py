"""Define owner service file."""
import uuid
from typing import Any, Dict, List, Union
from fastapi import FastAPI, Response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from io import BytesIO
import decouple
from fastapi import Depends
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from sqlalchemy.orm import Session

from app.src import models
from app.src.exceptions.error_code import AuthErrorCode, BEErrorCode
from app.src.models import Owner
from app.src.repositories.owner import OwnerRepository
from app.src.schemas.owner import OwnerCreate, OwnerUpdate
from app.src.repositories.group import GroupRepository

class OwnerService(object):
    """Define Owner service object."""

    def __init__(self) -> None:
        """Define constructor for Owner service object."""
        self.owner_repository = OwnerRepository(models.Owner)
        self.group_repository = GroupRepository(models.Group)

    def get_owner_by_id(self, db_session: Session, owner_id: uuid.UUID) -> models.Owner:
        """Define get owner by id method."""
        owner = self.owner_repository.get(db_session, obj_id=owner_id)
        if not owner:
           raise BEErrorCode.OWNER_NOT_FOUND.value
        return owner

    def create_owner(self, db_session: Session, owner_create: OwnerCreate,) -> models.Owner:
        """Define create owner method."""

        if owner_create.name=='string':
             del owner_create['name']
        if owner_create.address=='string':
             owner_create.address = None
        if owner_create.garbageMass==0:
             owner_create.garbageMass = 0
        if owner_create.neighbourhood_id=='string':
             owner_create.neighbourhood_id=None
        if owner_create.group_id=='string':
             owner_create.group_id=None
        if owner_create.state_id=='string':
             owner_create.state_id=None
        if owner_create.force_id=='string':
             owner_create.force_id=None
        if owner_create.group_id!='string':
            group_name = self.group_repository.get(db_session, owner_create.group_id).name
            if group_name == "Hộ gia đình":
                owner_create.price=312000
            if group_name == "Nhóm 1":
                owner_create.price=312000
            if group_name == "Nhóm 2":
                owner_create.price=480000
            if group_name == "Nhóm 3":
                owner_create.price=876000
            if group_name == "Nhóm 4":
                owner_create.price=owner_create.garbageMass*217500*12
        owner = self.owner_repository.create(db_session, obj_in=owner_create)
        return owner

    def delete_owner(self, db_session: Session, owner_id: uuid.UUID) -> None:
        """Define remove owner method."""
        owner = self.owner_repository.get(db_session, obj_id=owner_id)
        if not owner:
            raise BEErrorCode.OWNER_NOT_FOUND.value
        self.owner_repository.delete(db_session, obj_id=owner.id)

    def update_owner(self, db_session: Session, onwer_id: uuid.UUID, owner_update: OwnerUpdate) -> models.Owner:
        """Update an existing owner."""
        owner = self.owner_repository.get(db_session, obj_id=onwer_id)
        if self.owner_repository.get(db_session, obj_id=onwer_id) is None:
            raise BEErrorCode.OWNER_NOT_FOUND.value
        if owner_update.name=='string':
            owner_update.name = owner.name
        if owner_update.address=='string':
            owner_update.address = owner.address
        if owner_update.garbageMass==0:
            owner_update.garbageMass =  owner.garbageMass
        if owner_update.neighbourhood_id=='string':
            owner_update.neighbourhood_id = owner.neighbourhood_id
        if owner_update.group_id=='string':
            owner_update.group_id = owner.group_id
        if owner_update.state_id=='string':
            owner_update.state_id = owner.state_id
        if owner_update.force_id == 'string':
            owner_update.force_id = owner.force_id
        owner = self.owner_repository.update(db_session, obj_id=onwer_id, obj_in=owner_update)
        owner = self.owner_repository.get(db_session, onwer_id)
        if not owner:
            raise BEErrorCode.OWNER_NOT_FOUND.value
        return owner
    
    def search_owners(self, db_session: Session, filters: OwnerUpdate) -> List[models.Owner]:
        """Define method to search owner based on filters."""

        owers = self.owner_repository.search_owners(db_session, filters)
        sorted_owner = sorted(owers, key=lambda x: x.created_at)
        if not owers:
            raise BEErrorCode.OWNER_NOT_FOUND.value
        return sorted_owner
    
    def generate_excel(self, db_session: Session, filters: OwnerUpdate) -> tuple[bytes, str]:
        """Generates and returns Excel data and filename for owner report."""
        sorted_owners = self.search_owners(db_session, filters)  
        wb = Workbook()
        ws = wb.active
        headers = ["Name", "Address", "Khoi Luong Rac", "Gia Tien", "Khu Pho", "Nhom", "Tinh Trang", "Luc Luong"]
        ws.append(headers)
        for col in ws.iter_cols(min_row=1, max_row=20):
            for cell in col:
                cell.font = Font(bold=True) 
        for owner in sorted_owners:
            owner_data = [owner.name, owner.address, owner.garbageMass, owner.price, str(owner.neighbourhood_id), str(owner.group_id), str(owner.state_id), str(owner.force_id)]
            ws.append(owner_data)

        filename = "owner_report.xlsx" 
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file.getvalue(), filename
    